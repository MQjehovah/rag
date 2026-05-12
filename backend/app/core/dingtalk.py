import asyncio
import io
import logging
import zipfile
import xml.etree.ElementTree as ET
from typing import List, Dict, Any, Optional, Callable, Awaitable

from alibabacloud_dingtalk.wiki_2_0.client import Client as WikiClient
from alibabacloud_dingtalk.wiki_2_0 import models as wiki_models
from alibabacloud_dingtalk.storage_1_0.client import Client as StorageClient
from alibabacloud_dingtalk.storage_1_0 import models as storage_models
from alibabacloud_tea_openapi import models as open_api_models
from alibabacloud_tea_util import models as util_models
import httpx

from app.config import settings

logger = logging.getLogger(__name__)

API_CALL_INTERVAL = 0.15
MAX_RETRIES = 3

DOWNLOADABLE_EXTENSIONS = {
    "pdf", "docx", "doc", "xlsx", "xls", "pptx", "ppt",
    "txt", "csv", "md", "html", "json",
}

WIKI_EXTENSIONS = {"", "wiki", "mindmap", "note"}
SKIP_EXTENSIONS = {
    "mp4", "avi", "mov", "mkv", "mp3", "wav",
    "exe", "dll", "bin", "pak", "dat",
    "jpg", "jpeg", "png", "gif", "bmp",
    "zip", "rar", "7z", "tar", "gz",
    "apk", "ipa", "dlink",
}


class DingTalkClient:
    def __init__(self):
        config = open_api_models.Config()
        config.protocol = "https"
        config.region_id = "central"
        self.wiki = WikiClient(config)
        self.storage = StorageClient(config)
        self._access_token: str = ""
        self._http = httpx.AsyncClient(timeout=30.0)
        self._last_call = 0.0
        self._dentry_cache: Dict[str, Dict[str, str]] = {}
        self._on_progress: Optional[Callable] = None
        self._collected_count: int = 0

    async def _throttle(self):
        now = asyncio.get_event_loop().time()
        elapsed = now - self._last_call
        if elapsed < API_CALL_INTERVAL:
            await asyncio.sleep(API_CALL_INTERVAL - elapsed)
        self._last_call = asyncio.get_event_loop().time()

    async def _get_token(self) -> str:
        if self._access_token:
            return self._access_token
        r = await self._http.get(
            "https://oapi.dingtalk.com/gettoken",
            params={
                "appkey": settings.dingtalk_app_key,
                "appsecret": settings.dingtalk_app_secret,
            },
        )
        r.raise_for_status()
        data = r.json()
        if data.get("errcode") != 0:
            raise Exception(f"DingTalk auth failed: {data}")
        self._access_token = data["access_token"]
        return self._access_token

    def _runtime(self) -> util_models.RuntimeOptions:
        rt = util_models.RuntimeOptions()
        rt.read_timeout = 30000
        rt.connect_timeout = 10000
        return rt

    async def list_workspaces(self) -> List[Dict[str, Any]]:
        token = await self._get_token()
        headers = wiki_models.ListWorkspacesHeaders()
        headers.x_acs_dingtalk_access_token = token
        result = []

        req = wiki_models.ListWorkspacesRequest(
            max_results=50,
            order_by="VIEW_TIME_DESC",
            with_permission_role=False,
            operator_id=settings.dingtalk_operator_id,
        )
        await self._throttle()
        resp = await self.wiki.list_workspaces_with_options_async(
            req, headers, self._runtime()
        )
        body = resp.body
        workspaces = body.workspaces if body.workspaces else []
        for ws in workspaces:
            result.append({
                "id": getattr(ws, "workspace_id", ""),
                "name": getattr(ws, "name", ""),
                "root_node_id": getattr(ws, "root_node_id", ""),
            })

        return result

    async def list_nodes(self, parent_node_id: str) -> List[Dict[str, Any]]:
        token = await self._get_token()
        headers = wiki_models.ListNodesHeaders()
        headers.x_acs_dingtalk_access_token = token
        result = []
        next_token = None

        while True:
            req = wiki_models.ListNodesRequest(
                max_results=30,
                parent_node_id=parent_node_id,
                operator_id=settings.dingtalk_operator_id,
                with_permission_role=False,
            )
            if next_token:
                req.next_token = next_token

            await self._throttle()
            resp = await self.wiki.list_nodes_with_options_async(
                req, headers, self._runtime()
            )
            body = resp.body
            nodes = body.nodes if body.nodes else []
            for n in nodes:
                result.append({
                    "node_id": getattr(n, "node_id", ""),
                    "name": getattr(n, "name", ""),
                    "type": getattr(n, "type", ""),
                    "extension": getattr(n, "extension", "") or "",
                    "category": getattr(n, "category", "") or "",
                    "has_children": getattr(n, "has_children", False),
                })

            next_token = getattr(body, "next_token", None) or ""
            if not next_token or not nodes:
                break
            await asyncio.sleep(0.2)

        return result

    async def _query_dentry_id(self, dentry_uuid: str) -> Optional[Dict[str, str]]:
        if dentry_uuid in self._dentry_cache:
            return self._dentry_cache[dentry_uuid]

        token = await self._get_token()
        for attempt in range(MAX_RETRIES):
            try:
                await self._throttle()
                r = await self._http.get(
                    f"https://api.dingtalk.com/v2.0/doc/dentries/{dentry_uuid}/queryDentryId",
                    headers={
                        "x-acs-dingtalk-access-token": token,
                    },
                    params={"operatorId": settings.dingtalk_operator_id},
                )
                r.raise_for_status()
                data = r.json()
                result = {
                    "space_id": data["spaceId"],
                    "dentry_id": data["dentryId"],
                }
                self._dentry_cache[dentry_uuid] = result
                return result
            except Exception as e:
                if "QpsLimit" in str(e) and attempt < MAX_RETRIES - 1:
                    await asyncio.sleep(2.0 * (attempt + 1))
                    continue
                logger.warning(f"queryDentryId failed for {dentry_uuid}: {e}")
                return None

    async def _download_file(self, space_id: str, dentry_id: str) -> Optional[bytes]:
        token = await self._get_token()
        for attempt in range(MAX_RETRIES):
            try:
                await self._throttle()
                sh = storage_models.GetFileDownloadInfoHeaders()
                sh.x_acs_dingtalk_access_token = token
                option = storage_models.GetFileDownloadInfoRequestOption(prefer_intranet=False)
                req = storage_models.GetFileDownloadInfoRequest(
                    union_id=settings.dingtalk_operator_id,
                    option=option,
                )
                resp = await self.storage.get_file_download_info_with_options_async(
                    space_id, dentry_id, req, sh, self._runtime()
                )
                sig = resp.body.header_signature_info
                urls = sig.resource_urls if sig.resource_urls else []
                hdrs = dict(sig.headers) if sig.headers else {}
                if not urls:
                    return None

                r = await self._http.get(urls[0], headers=hdrs, follow_redirects=True)
                r.raise_for_status()
                return r.content
            except Exception as e:
                if "QpsLimit" in str(e) and attempt < MAX_RETRIES - 1:
                    await asyncio.sleep(2.0 * (attempt + 1))
                    continue
                logger.warning(f"Download failed {space_id}/{dentry_id}: {e}")
                return None

    @staticmethod
    def _extract_text(content: bytes, extension: str) -> str:
        ext = extension.lower()
        if ext == "docx":
            try:
                return DingTalkClient._docx_to_markdown(content)
            except Exception:
                pass
        if ext == "xlsx":
            try:
                from openpyxl import load_workbook
                wb = load_workbook(io.BytesIO(content), read_only=True)
                rows = []
                for ws in wb.worksheets:
                    for row in ws.iter_rows(values_only=True):
                        cells = [str(c) for c in row if c is not None]
                        if cells:
                            rows.append("| " + " | ".join(cells) + " |")
                wb.close()
                return "\n".join(rows)
            except Exception:
                pass
        if ext == "pptx":
            try:
                buf = io.BytesIO(content)
                with zipfile.ZipFile(buf) as z:
                    slides = []
                    slide_files = sorted([n for n in z.namelist() if n.startswith("ppt/slides/slide") and n.endswith(".xml")])
                    for name in slide_files:
                        with z.open(name) as f:
                            tree = ET.parse(f)
                            ns = "http://schemas.openxmlformats.org/drawingml/2006/main"
                            texts = [t.text for t in tree.getroot().iter(f"{{{ns}}}t") if t.text]
                            if texts:
                                slides.append(" ".join(texts))
                    return "\n\n---\n\n".join(slides)
            except Exception:
                pass
        try:
            return content.decode("utf-8", errors="ignore")
        except Exception:
            return ""

    @staticmethod
    def _docx_to_markdown(content: bytes) -> str:
        buf = io.BytesIO(content)
        lines = []
        with zipfile.ZipFile(buf) as z:
            with z.open("word/document.xml") as f:
                tree = ET.parse(f)
                root = tree.getroot()
                ns = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
                ns_p = ns
                ns_r = ns

                for para in root.iter(f"{{{ns_p}}}p"):
                    style_el = para.find(f".//{{{ns_p}}}pStyle")
                    style = style_el.get(f"{{{ns_p}}}val") if style_el is not None else ""

                    runs = []
                    for run in para.iter(f"{{{ns_r}}}r"):
                        rpr = run.find(f"{{{ns_r}}}rPr")
                        bold = rpr is not None and rpr.find(f"{{{ns_r}}}b") is not None
                        italic = rpr is not None and rpr.find(f"{{{ns_r}}}i") is not None
                        text_els = run.findall(f"{{{ns_r}}}t")
                        text = "".join(t.text or "" for t in text_els)
                        if not text:
                            continue
                        if bold and italic:
                            text = f"***{text}***"
                        elif bold:
                            text = f"**{text}**"
                        elif italic:
                            text = f"*{text}*"
                        runs.append(text)

                    line = "".join(runs)
                    if not line.strip():
                        lines.append("")
                        continue

                    num_el = para.find(f".//{{{ns_p}}}numPr")
                    if num_el is not None:
                        lines.append(f"- {line}")
                    elif "Heading1" in style or style == "1":
                        lines.append(f"# {line}")
                    elif "Heading2" in style or style == "2":
                        lines.append(f"## {line}")
                    elif "Heading3" in style or style == "3":
                        lines.append(f"### {line}")
                    elif "Heading4" in style or style == "4":
                        lines.append(f"#### {line}")
                    elif "Title" in style:
                        lines.append(f"# {line}")
                    else:
                        lines.append(line)

        return "\n\n".join(lines)

    async def download_node_content(self, node_id: str, extension: str) -> Optional[str]:
        dentry = await self._query_dentry_id(node_id)
        if not dentry:
            return None
        data = await self._download_file(dentry["space_id"], dentry["dentry_id"])
        if not data:
            return None
        if not extension:
            return data.decode("utf-8", errors="ignore")
        return self._extract_text(data, extension)

    async def collect_all_docs(
        self, workspace_id: str = None,
        on_progress: Optional[Callable[[Dict[str, Any], int], Awaitable[None]]] = None,
    ) -> List[Dict[str, Any]]:
        target_id = workspace_id or settings.dingtalk_knowledge_base_id
        logger.info(f"Starting DingTalk sync, workspace_id={target_id or 'all'}")

        self._on_progress = on_progress
        self._collected_count = 0

        spaces = await self.list_workspaces()

        if target_id:
            spaces = [s for s in spaces if s["id"] == target_id]
            if not spaces:
                logger.error(f"Workspace {target_id} not found")
                return []

        logger.info(f"Found {len(spaces)} workspaces")

        all_docs = []
        for sp in spaces:
            sp_id = sp["id"]
            sp_name = sp["name"]
            root_id = sp.get("root_node_id") or sp_id
            logger.info(f"Crawling workspace: {sp_name} ({sp_id}) root={root_id}")

            try:
                docs = await self._crawl_recursive(sp_name, root_id)
                all_docs.extend(docs)
            except Exception as e:
                logger.error(f"Failed to crawl workspace {sp_name}: {e}")

        return all_docs

    async def _crawl_recursive(
        self, workspace_name: str, parent_id: str, path: str = ""
    ) -> List[Dict[str, Any]]:
        docs = []
        try:
            nodes = await self.list_nodes(parent_id)
        except Exception as e:
            logger.error(f"Failed to list nodes: {e}")
            return docs

        for node in nodes:
            node_id = node["node_id"]
            node_type = node["type"]
            name = node["name"] or "无标题"
            ext = node.get("extension", "")
            has_children = node.get("has_children", False)
            current_path = f"{path}/{name}" if path else name

            if node_type == "FOLDER" or has_children:
                child_docs = await self._crawl_recursive(
                    workspace_name, node_id, current_path
                )
                docs.extend(child_docs)
                continue

            if ext.lower() in SKIP_EXTENSIONS:
                logger.debug(f"  Skipped: {name} (ext={ext})")
                continue

            content = None
            if ext.lower() in DOWNLOADABLE_EXTENSIONS:
                content = await self.download_node_content(node_id, ext)
            elif ext.lower() in WIKI_EXTENSIONS or node_type == "DOC":
                content = await self.download_node_content(node_id, "docx")
                if not content:
                    content = await self.download_node_content(node_id, "")
            else:
                logger.debug(f"  Skipped: {name} (unsupported ext={ext})")
                continue

            if content and content.strip():
                doc = {
                    "id": node_id,
                    "title": name,
                    "content": content,
                    "space_name": workspace_name,
                    "path": current_path,
                }
                docs.append(doc)
                self._collected_count += 1
                logger.info(f"  Collected: {name} ({len(content)} chars)")
                if self._on_progress:
                    try:
                        await self._on_progress(doc, self._collected_count)
                    except Exception:
                        pass
            else:
                logger.warning(f"  Empty content: {name}")

        return docs

    async def close(self):
        await self._http.aclose()
